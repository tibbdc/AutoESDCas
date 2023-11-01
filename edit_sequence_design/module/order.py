
import xlsxwriter
import time  
import sys,os
import math
from sgRNA_utils.sgRNA_primer_config import config 
import pandas as pd

# 自定义单元格格式
def set_merge_style(workbook, tag):
#     bold,font_size,border,align,font_color='','','','',''
    if tag == 'head1':
        bold, font_size, border = True, 20, 1
        font_color = 'black'
        align = 'center'
        num_format = '0.00'
    elif tag == 'cell1':
        bold, font_size, border = False, 12, 1
        border = 1
        font_color = 'black'
        align = 'left'
        num_format='0.00'
    elif tag == 'date':
        num_format = 'd/mmmm/yyyy'
        bold, font_size, border = False, 12, 1
        font_color = 'black'
        align = 'left'
        
    elif tag == 'cell2':
        bold, font_size, border = False, 12, 1
        font_color = 'black'
        align = 'center'
        num_format='0.00'
    elif tag == 'url':
        bold, font_size, border = False, 12, 1
        font_color = 'blue'
        align = 'center'
        num_format='0.00'
    else:
        bold, font_size, border = False, 12, 1
        font_color = 'red'
        align = 'left'
        num_format='0.00'
    
    cell_format = workbook.add_format({
        'bold': bold,
        'font_size': font_size,
        'border': border,
        'align': align,
        'valign': 'vcenter',  # 垂直居中
        'font_color': font_color,
        'num_format':num_format,  
        'font_name': '微软雅黑',
        'text_wrap':True     
    }) 
    return cell_format   

# 输出引物订单
def create_orders(primer_orders_df,orders_path=config.OUTPUT_FILE_PATH+config.OUTPUT_ORDERS_NAME):
    workbook = xlsxwriter.Workbook(orders_path)
    worksheet=workbook.add_worksheet('常规引物合')
    #设置行高
    worksheet.set_row(0, 70)
    worksheet.set_row(1, 25)   
    worksheet.set_row(2, 25)
    worksheet.set_row(3, 25)
    worksheet.set_row(4, 25)
    worksheet.set_row(5, 25)
    worksheet.set_row(6, 25)
    worksheet.set_row(7, 25)
    worksheet.set_row(8, 30)
    #设置列宽
    worksheet.set_column('A:A', 25)
    worksheet.set_column('B:B', 35)
    worksheet.set_column('C:C', 40)
    worksheet.set_column('D:D', 6)
    worksheet.set_column('E:E', 15)
    worksheet.set_column('H:H', 15)
    worksheet.set_column('I:I', 20)
    #插入图片
    print( config.DATA_ROOT +config.INPUT_IMG_PATH + config.IMG_ORDERS_NAME )   
    worksheet.insert_image('A1', config.DATA_ROOT +config.INPUT_IMG_PATH + config.IMG_ORDERS_NAME  ,{'x_offset': 15, 'y_offset': 10})

    
    worksheet.merge_range('A1:I1', '中国科学院天津工业生物技术研究所引物合成订购单',set_merge_style(workbook,tag='head1'))  
    worksheet.merge_range('B2:C2',time.strftime('%Y-%m-%d'),set_merge_style(workbook,tag='date'))
    worksheet.merge_range('B3:C3','',set_merge_style(workbook,tag='cell1'))
    worksheet.merge_range('B4:C4','',set_merge_style(workbook,tag='cell2')) 
    worksheet.merge_range('B5:C5','中国科学院天津工业生物技术研究所',set_merge_style(workbook,tag='cell1')) 
    worksheet.merge_range('B6:C6','空港经济区西七道32号',set_merge_style(workbook,tag='cell1'))
    worksheet.merge_range('B7:C7','',set_merge_style(workbook,tag='cell1'))  
    worksheet.write_string('B7','',set_merge_style(workbook,tag='cell1'))

    worksheet.merge_range('B8:C8','', set_merge_style(workbook, tag='cell2'))  
    worksheet.write_url('B8', '', set_merge_style(workbook, tag='url'), '')
    worksheet.write_string('A2','订购日期',set_merge_style(workbook, tag='cell2'))

    # color = workbook.add_format({'color': 'red'})


    col = 0
    for row, string_parts in enumerate([
        
        [set_merge_style(workbook,tag='other'), '*', set_merge_style(workbook,tag='cell2'), '客户姓名'],
        [set_merge_style(workbook,tag='other'), '*', set_merge_style(workbook,tag='cell2'), '所属课题组'],
        [set_merge_style(workbook,tag='other'), '*', set_merge_style(workbook,tag='cell2'), '客户单位'],
        [set_merge_style(workbook,tag='other'), '*', set_merge_style(workbook,tag='cell2'), '客户地址'],
        [set_merge_style(workbook,tag='other'), '*', set_merge_style(workbook,tag='cell2'), '客户电话'],  
        [set_merge_style(workbook,tag='other'), '*', set_merge_style(workbook,tag='cell2'), '客户Email'],
        
    ]):
        print(string_parts)
        worksheet.write_rich_string(row+2, col, *string_parts,set_merge_style(workbook,tag='cell2'))

    worksheet.merge_range('D2:G2','填写要求',set_merge_style(workbook,tag='cell2'))
    worksheet.merge_range('H2:I2','常用兼并碱基代码',set_merge_style(workbook,tag='cell2'))

    worksheet.merge_range('D3:G8','', set_merge_style(workbook,tag='cell1')) 
    worksheet.write_rich_string('D3',     
                                set_merge_style(workbook,tag='cell1'), '1.客户信息栏中*表示必填项；\n2.点击表格正文中的空行可以查看填写说明；\n3.订单发送前请仔细核对信息, 发送后将不能更改。\n', 
                                set_merge_style(workbook,tag='other'), '请上传至  www.biosys.........cn。\n',
                                set_merge_style(workbook,tag='cell1'), '4.服务电话:022-24828729',
                                set_merge_style(workbook,tag='cell1'))

    worksheet.write_string('H3','M:(A/C)',set_merge_style(workbook,tag='cell1'))
    worksheet.write_string('H4','R:(A/G)',set_merge_style(workbook,tag='cell1'))
    worksheet.write_string('H5','W:(A/T)',set_merge_style(workbook,tag='cell1'))
    worksheet.write_string('H6','S:(G/C)',set_merge_style(workbook,tag='cell1'))
    worksheet.write_string('H7','Y:(C/T)',set_merge_style(workbook,tag='cell1'))
    worksheet.write_string('H8','K:(G/T)',set_merge_style(workbook,tag='cell1'))

    worksheet.write_string('I3','V:(A/G/C)',set_merge_style(workbook,tag='cell1'))
    worksheet.write_string('I4','H:(A/C/T)',set_merge_style(workbook,tag='cell1'))
    worksheet.write_string('I5','D:(A/G/T)',set_merge_style(workbook,tag='cell1'))
    worksheet.write_string('I6','B:(G/C/T)',set_merge_style(workbook,tag='cell1'))
    worksheet.write_string('I7','N:(A/G/C/T)',set_merge_style(workbook,tag='cell1'))

    worksheet.write_string('A9','板号',set_merge_style(workbook,tag='cell2'))
    worksheet.write_string('B9','孔的位置',set_merge_style(workbook,tag='cell2'))
    worksheet.write_string('C9',"碱基序列(5'to3')",set_merge_style(workbook,tag='cell2'))  
    worksheet.write_string('D9','碱基数',set_merge_style(workbook,tag='cell2'))  
    worksheet.write_string('E9','纯化方法',set_merge_style(workbook,tag='cell2'))
    worksheet.write_string('F9','总nmol数',set_merge_style(workbook,tag='cell2')) 
    worksheet.write_string('G9','分装管数',set_merge_style(workbook,tag='cell2'))
    worksheet.write_string('H9','总OD值',set_merge_style(workbook,tag='cell2'))
    worksheet.write_string('I9','备注',set_merge_style(workbook,tag='cell2'))
    # worksheet.merge_range('I9','备注',set_merge_style(workbook,tag='cell2'))
       
    # for i in range(10,len(primer_orders_df)+11):
    #     worksheet.merge_range(f'H{i}:I{i}','')   
  

    df_length = len(primer_orders_df)
    worksheet.add_table(f'A10:D{df_length+10}',{'data':primer_orders_df.values.tolist(),
                                            'header_row': False,  
                                            'style':None
                                            })
    workbook.close()

#
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment,Side,Border,PatternFill


def set_cell_font_style(file, sheetName, rownum, columno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    
    sheet.cell(rownum, columno).alignment = Alignment(
    horizontal='center',     
    vertical='center',        
    text_rotation=0,       
    wrap_text=False,       
    shrink_to_fit=False,   
    indent=0,              
    )
    side = Side(
    style="medium",  
    color="FFFFFF",  
    )
    
    sheet.cell(rownum, columno).border = Border(
    top=side,   
    bottom=side, 
    left=side,   
    right=side,   
    diagonal=side
    )
    workbook.save(file)
    
    
def fillGreenColor(file, sheetName, rownum, columno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    greenFill = PatternFill(start_color='60b212',
                            end_color='60b212',
                            fill_type='solid')
    sheet.cell(rownum, columno).fill = greenFill
    workbook.save(file)


def writeData(file, sheetName, rowno, colno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rowno, colno).value = data
    workbook.save(file)


def Rewrite_excel_template(file,unique_all_primer_df,start_row,start_col,sheetName):

    file = "/home/yanghe/githubcode/crispr_hr_editor/edit_sequence_design/data/input/123.xlsx"
    sheetName = '引物合成订单'
    start_row = 13
    start_col = 2
    for i in range(len(unique_all_primer_df)):
        for j in range(len(unique_all_primer_df.columns)):
            writeData(file, sheetName, start_row+i, start_col+j, unique_all_primer_df.iloc[i,j])
            set_cell_font_style(file, sheetName, start_row+i, start_col+j)

#重命名id
def rename_primer_ID(temp_df,in_col,df_name,stype):
    temp_df = temp_df[in_col]
    temp_df.index = temp_df['ID']
    temp_df = temp_df.drop(columns='ID')
    temp_v = pd.DataFrame(temp_df.unstack())
    if stype == 1:
        temp_v.index = [str(i[1])+'|'+ df_name + '|'+ str(i[0][7])+'|' for i in temp_v.index]
    elif stype == 2:
        temp_v.index = [str(i[1])+'|'+ df_name + '|'+ str(i[0].split('_')[-1]) +'|' for i in temp_v.index]
    return temp_v

# #引物重命名
# def order_primer_rename_by_ways(primer_orders_df,label=['A','B','C','D','E','F','G','H']):    
#     primer_name_df = pd.DataFrame()
#     cell = math.ceil(len(primer_orders_df) / len(label))
#     cell_remainder = len(primer_orders_df) % len(label)
#     big_cell = math.ceil( cell/12 )
#     big_cell_remainder = cell % 12

#     for i in range(1,big_cell+1):
#         for j in range(1,13):
#             primer_name_df =primer_name_df.append(pd.DataFrame(label)+str(j))
            
#     #重置索引        
#     primer_name_df.reset_index(drop=True,inplace=True)

#     #减去余数    
#     if big_cell_remainder !=0:
#         primer_name_df=primer_name_df.iloc[:-(big_cell_remainder*12*8),:]
#         for j in range(1,big_cell_remainder+1):
#             primer_name_df =primer_name_df.append(pd.DataFrame(label)+str(j))

#     if cell_remainder != 0:
#         primer_name_df=primer_name_df.iloc[:-(len(label)-cell_remainder),:]
        
#     name = primer_name_df[0].values.tolist()
# #     primer_name_df.rename(columns={0:'引物名称'},inplace=True)
#     return name


#引物重命名
def order_primer_rename_by_ways(primer_orders_df,label=['A','B','C','D','E','F','G','H']):    
    primer_name_df = pd.DataFrame()
    cell = math.floor(len(primer_orders_df) / len(label))
    cell_remainder = len(primer_orders_df) % len(label)
    big_cell = math.floor( cell/12 )
    big_cell_remainder = cell % 12
    print(cell,cell_remainder,big_cell,big_cell_remainder)
    j = 1
    for i in range(1,cell+1):
        if j > 12:
            j = 1
        primer_name_df =primer_name_df.append(pd.DataFrame(label)+str(j))
        j=j+1
    primer_name_df = primer_name_df.append(pd.DataFrame(label[:cell_remainder])+str(j))
    name = primer_name_df[0].values.tolist()
    primer_name_df.rename(columns={0:'引物名称'},inplace=True)

    return name



#引物合并
def merge_primer(*li):  
    if len(li) == 7:
        uha_primer_df,dha_primer_df,sgRNA_primer_df,plasmid_backbone_p_df,seq_altered_p_df,plasmid_sequencing_primer_df,genome_sequencing_primer_df = li
        #给质粒骨架引物的id命名
        plasmid_backbone_p_df['ID'] = 'plasmid'
    elif len(li) == 8:
        uha_primer_df,dha_primer_df,sgRNA_primer_df,ccdb_primer_df,seq_altered_p_df,sgRNA_plasmid_sequencing_primer_df,ccdb_plasmid_sequencing_primer_df,genome_sequencing_primer_df = li
    elif len(li) == 2:
        plasmid_sequencing_primer_df,genome_sequencing_primer_df = li
        uha_primer_df = pd.DataFrame()
        dha_primer_df = pd.DataFrame()
        sgRNA_primer_df = pd.DataFrame()
        plasmid_backbone_p_df = pd.DataFrame()
        seq_altered_p_df = pd.DataFrame()
    elif len(li) == 3:
        sgRNA_plasmid_sequencing_primer_df,ccdb_plasmid_sequencing_primer_df,genome_sequencing_primer_df = li
        uha_primer_df = pd.DataFrame()
        dha_primer_df = pd.DataFrame()
        sgRNA_primer_df = pd.DataFrame()
        plasmid_backbone_p_df = pd.DataFrame()
        seq_altered_p_df = pd.DataFrame()
    elif len(li) == 1:
        genome_sequencing_primer_df = li[0]
        uha_primer_df = pd.DataFrame()
        dha_primer_df = pd.DataFrame()
        sgRNA_primer_df = pd.DataFrame()
        plasmid_backbone_p_df = pd.DataFrame()
        seq_altered_p_df = pd.DataFrame()
        plasmid_sequencing_primer_df = pd.DataFrame()



    u_primer = pd.DataFrame()
    d_primer = pd.DataFrame()
    p_s_primer = pd.DataFrame()
    g_s_primer = pd.DataFrame()
    sgRNA_primer = pd.DataFrame()
    plasmid_backbone_primer = pd.DataFrame()


    #同源臂、sgRNA
     
    if len(uha_primer_df) > 0:
        u_primer = rename_primer_ID(uha_primer_df,in_col=['ID','PRIMER_LEFT_WHOLE_SEQUENCE','PRIMER_RIGHT_WHOLE_SEQUENCE'],df_name='u',stype=1)
    if len(dha_primer_df) > 0:
        d_primer = rename_primer_ID(dha_primer_df,in_col=['ID','PRIMER_LEFT_WHOLE_SEQUENCE','PRIMER_RIGHT_WHOLE_SEQUENCE'],df_name='d',stype=1)
    if len(seq_altered_p_df) > 0:
        seq_altered_primer = rename_primer_ID(seq_altered_p_df,in_col=['ID','PRIMER_LEFT_WHOLE_SEQUENCE','PRIMER_RIGHT_WHOLE_SEQUENCE'],df_name='s',stype=1)

    #基因组测序
    if len(genome_sequencing_primer_df)>0:
        genome_sequencing_primer_df_columns = [i for i in genome_sequencing_primer_df.columns if (i.split('_')[-1]).isdigit() ]
        # genome_sequencing_primer_df_columns = [i for i in genome_sequencing_primer_df.columns ]
        genome_sequencing_primer_df_columns.append('ID')
        g_s_primer = rename_primer_ID(genome_sequencing_primer_df, in_col=genome_sequencing_primer_df_columns, df_name='g', stype=2)


    if len(li) == 7:
        if len(sgRNA_primer_df) > 0 :
            sgRNA_primer = rename_primer_ID(sgRNA_primer_df,in_col=['ID','PRIMER_LEFT_WHOLE_SEQUENCE','PRIMER_RIGHT_WHOLE_SEQUENCE'],df_name='f',stype=1)
        if len(plasmid_backbone_p_df) > 0:
            plasmid_backbone_primer = rename_primer_ID(plasmid_backbone_p_df,in_col=['ID','PRIMER_LEFT_WHOLE_SEQUENCE','PRIMER_RIGHT_WHOLE_SEQUENCE'],df_name='b',stype=1)

        #质粒测序
        
        if len(plasmid_sequencing_primer_df) >0:
            plasmid_sequencing_primer_df_columns = [i for i in plasmid_sequencing_primer_df.columns if (i.split('_')[-1]).isdigit() ]
            plasmid_sequencing_primer_df_columns.append('ID')
            p_s_primer = rename_primer_ID(plasmid_sequencing_primer_df,in_col=plasmid_sequencing_primer_df_columns,df_name='p',stype=2)

        #合并引物
        if len(seq_altered_p_df)> 0:
            all_primer_df = pd.concat([u_primer,d_primer,sgRNA_primer,plasmid_backbone_primer,seq_altered_primer,p_s_primer,g_s_primer])
        else:
            all_primer_df = pd.concat([u_primer,d_primer, sgRNA_primer, plasmid_backbone_primer, p_s_primer, g_s_primer])

        
    elif len(li)==8:

        if len(sgRNA_primer_df) >0:
            if len(sgRNA_primer_df.columns) < 4:
                sgRNA_primer = rename_primer_ID(sgRNA_primer_df,in_col=['ID','Target sequence'],df_name='sgRNA',stype=1)
            else:  
                sgRNA_primer = rename_primer_ID(sgRNA_primer_df,in_col=['ID','PRIMER_LEFT_WHOLE_SEQUENCE','PRIMER_RIGHT_WHOLE_SEQUENCE'],df_name='sgRNA',stype=1)
        if len(ccdb_primer_df) >0:
            ccdb_primer = rename_primer_ID(ccdb_primer_df,in_col=['ID','PRIMER_LEFT_WHOLE_SEQUENCE','PRIMER_RIGHT_WHOLE_SEQUENCE'],df_name='ccdb',stype=1)
        else:
            ccdb_primer = pd.DataFrame()

        #sgRNA质粒测序引物
        if len(sgRNA_plasmid_sequencing_primer_df) > 0 :
            sgRNA_plasmid_sequencing_primer_df_columns = [i for i in sgRNA_plasmid_sequencing_primer_df.columns if (i.split('_')[-1]).isdigit() ]
            sgRNA_plasmid_sequencing_primer_df_columns.append('ID')
            p1_s_primer = rename_primer_ID(sgRNA_plasmid_sequencing_primer_df, in_col=sgRNA_plasmid_sequencing_primer_df_columns, df_name='p1',stype=2)
        else:
            p1_s_primer = pd.DataFrame()
        #ccdb质粒测序引物
        if len(ccdb_plasmid_sequencing_primer_df):
            ccdb_plasmid_sequencing_primer_df_columns = [i for i in ccdb_plasmid_sequencing_primer_df.columns if (i.split('_')[-1]).isdigit() ]
            ccdb_plasmid_sequencing_primer_df_columns.append('ID')
            p2_s_primer = rename_primer_ID(ccdb_plasmid_sequencing_primer_df, in_col=ccdb_plasmid_sequencing_primer_df_columns, df_name='p2',stype=2)
        else:
            p2_s_primer = pd.DataFrame()  

        #合并引物   
        if len(seq_altered_p_df)> 0:
            all_primer_df = pd.concat([u_primer,d_primer,sgRNA_primer,ccdb_primer,seq_altered_primer,p1_s_primer,p2_s_primer,g_s_primer])
        else:
            all_primer_df = pd.concat([u_primer,d_primer,sgRNA_primer,ccdb_primer,p1_s_primer,p2_s_primer,g_s_primer])


    elif len(li) == 2:
         #质粒测序
        if len(plasmid_sequencing_primer_df) >0:
            plasmid_sequencing_primer_df_columns = [i for i in plasmid_sequencing_primer_df.columns if (i.split('_')[-1]).isdigit() ]
            plasmid_sequencing_primer_df_columns.append('ID')
            p_s_primer = rename_primer_ID(plasmid_sequencing_primer_df,in_col=plasmid_sequencing_primer_df_columns,df_name='p',stype=2)

        #合并引物
        if len(seq_altered_p_df)> 0:
            all_primer_df = pd.concat([u_primer,d_primer,sgRNA_primer,plasmid_backbone_primer,seq_altered_primer,p_s_primer,g_s_primer])
        else:
            all_primer_df = pd.concat([u_primer,d_primer, sgRNA_primer, plasmid_backbone_primer, p_s_primer, g_s_primer])
    elif len(li) == 1:
        all_primer_df = g_s_primer
    
    elif len(li) == 3:

        #sgRNA质粒测序引物
        if len(sgRNA_plasmid_sequencing_primer_df) > 0 :
            sgRNA_plasmid_sequencing_primer_df_columns = [i for i in sgRNA_plasmid_sequencing_primer_df.columns if (i.split('_')[-1]).isdigit() ]
            sgRNA_plasmid_sequencing_primer_df_columns.append('ID')
            p1_s_primer = rename_primer_ID(sgRNA_plasmid_sequencing_primer_df, in_col=sgRNA_plasmid_sequencing_primer_df_columns, df_name='p1',stype=2)
        else:
            p1_s_primer = pd.DataFrame()
        #ccdb质粒测序引物
        if len(ccdb_plasmid_sequencing_primer_df):
            ccdb_plasmid_sequencing_primer_df_columns = [i for i in ccdb_plasmid_sequencing_primer_df.columns if (i.split('_')[-1]).isdigit() ]
            ccdb_plasmid_sequencing_primer_df_columns.append('ID')
            p2_s_primer = rename_primer_ID(ccdb_plasmid_sequencing_primer_df, in_col=ccdb_plasmid_sequencing_primer_df_columns, df_name='p2',stype=2)
        else:
            p2_s_primer = pd.DataFrame()  
        
        #合并引物   
        all_primer_df = pd.concat([p1_s_primer, p2_s_primer, g_s_primer])



    all_primer_df = all_primer_df.drop_duplicates(keep='first', subset=[0])
    all_primer_df =  all_primer_df.dropna()
    
      
    
    all_primer_df = all_primer_df[all_primer_df[0]!=''] 
    #排序
    unique_all_primer_df = all_primer_df.reset_index().rename(columns={0:'primer'})
    

    unique_all_primer_df_hole_name = order_primer_rename_by_ways(unique_all_primer_df)

    
    unique_all_primer_df['index'] = [f'M{math.ceil(i/96)}|{v}' for i,v in zip(range(1,len(all_primer_df.index)+1),all_primer_df.index)]

    unique_all_primer_df['board_num'] = unique_all_primer_df['index'].apply(lambda x: x.split('|')[0].replace('M','Eco_101_gene_del_primer_'))



    unique_all_primer_df.insert(1,column='hole',value=unique_all_primer_df_hole_name)
    def work(x):
        if str(x) == 'nan':
            return 0
        else:
            return len(x)
    unique_all_primer_df['base_num'] = unique_all_primer_df.primer.apply(lambda x: work(x))


    


    return unique_all_primer_df


